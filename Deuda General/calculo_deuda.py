from datetime import datetime ; year = datetime.now().year ; salida = f'Finalistas-Leasing{year}.xlsx'
import math, pandas as pd

# Funcion que trunca el número
def truncar(numero):
    factor = 10 ** 2
    return math.trunc(numero * factor) / factor

# Funcion que obtiene el precio de cta.
def cuota_leasing(Precio, Porc_Beneficio, Residuo):
    descto = 0.5 - 0.25 *truncar(Porc_Beneficio)
    numero = ((Precio * descto) - Residuo)/10
    return round(numero)

# Obtener valor del Precio:
precios = open("Precios.txt", "r")
valor = residuo = 0
for linea in precios:
    linea = linea.strip()
    if (linea[0] == 'M'):        
        valor = int(linea.split('$')[-1])
    elif (linea[0]== 'R'):
        residuo = int(linea.split('$')[-1])
precios.close()


# Cargar el archivo Excel de 'alumnos.xlsx'
df = pd.read_excel('alumnos.xlsx')
df_finalistas = df[(df['LEASING HISTORICO'] == 'No') & (df['POSTULO LEASING 2024'] == 'Sí')]
df_finalistas['BENEFICIO RREE'] = pd.to_numeric(df_finalistas['BENEFICIO RREE'], errors='coerce')

# Calcular Cuota_Leasing usando la función cuota_leasing
df_finalistas['Cuota_Leasing'] = df_finalistas['BENEFICIO RREE'].apply(lambda x: cuota_leasing(valor, x, residuo))


# Añadir la columna 'OC' con el valor del 'residuo'
df_finalistas['OC'] = residuo
# Añadir la columna 'Total a Pagar' que sigue la fórmula (Cuota_Leasing * 10) + residuo
df_finalistas['Total a Pagar'] = (df_finalistas['Cuota_Leasing'] * 10) + residuo


df_finalistas = df_finalistas[['NOMBRES', 'PATERNO', 'MATERNO', 'RUT_DV', 'NOM_CARRERA', 'NOMBRE_SEDE', 'EMAIL_USM', 'BENEFICIO RREE', 'Cuota_Leasing', 'OC', 'Total a Pagar']]

# Guardar los finalistas en un nuevo archivo Excel 'Finalistas.xlsx'
df_finalistas.to_excel(salida, index=False)


from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
wb = load_workbook(salida)
ws = wb.active

# Definir el estilo para la primera fila (encabezados)
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center")

# Definir bordes para todas las celdas
thin_border = Border(
    left=Side(style="thin"), 
    right=Side(style="thin"), 
    top=Side(style="thin"), 
    bottom=Side(style="thin")
)

# Aplicar formato a todas las celdas (bordes)
for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border

# Aplicar formato especial a la primera fila (encabezado)
for cell in ws[1]:  # Fila 1 (encabezado)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
wb.save(salida)
    

import tkinter as tk
from tkinter import messagebox
root = tk.Tk()
root.withdraw()  # Ocultar la ventana principal de Tkinter
messagebox.showinfo("Proceso completado", f"Archivo '{salida}' creado con éxito.")
root.quit()